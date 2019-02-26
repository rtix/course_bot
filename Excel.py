import xlsxwriter
import openpyxl
import Database.CourseDB as CourseDB
import Database.UserDB as UserDB
import os.path

if not os.path.exists('ExcelData'):
    os.makedirs('ExcelData')

class ExtensionException(Exception):
	def __init__(self, value):
		self.value = value
	def __str__(self):
		return str('{0} could not be read as excel file'.format(self.value))

class CellValueException(Exception):
	def __init__(self, row, col, message):
		self.row = row
		self.col = col
		self.message = message
	def __str__(self):
		return str('Exception in {0}x{1} cell.\n{2}'.format(self.row, self.col, self.message))

def get_attendance(course_id):
	file_name = 'ExcelData/Attendance_{0}.xlsx'.format(course_id)
	book = xlsxwriter.Workbook(file_name)
	sheet = book.add_worksheet('Посещаемость')
	h_sheet = book.add_worksheet('Hidden')
	h_sheet.hide()

	course = CourseDB.get_course(course_id)

	centered_format = book.add_format(
									{
										'align': 'center',
                                   		'valign': 'vcenter',
                                   		'border': 1
                               		})
	users = list([UserDB.get_user(u_id) for u_id in course['participants']])
	sheet.merge_range(0, 0, 1, 0, 'Студенты', centered_format)
	sheet.set_column(0, 0, len(max([user['name'] for user in users])) + 4)

	cw_count = len(course['cw_numbers'])
	sheet.merge_range(0, 1, 1, cw_count, 'Посещаемость', centered_format)
	for i, class_work in enumerate([CourseDB.get_classwork(course_id, c_id) for c_id in course['cw_numbers']]):
		sheet.write(2, i+1, class_work['date'])
		h_sheet.write(2, i+1, class_work['classwork_id'])

	for i, user in enumerate(users):
		sheet.write(i+3, 0, str(i+1) + '. ' + user['name'])
		h_sheet.write(i+3, 0, user['u_id'])

		for j, attendance in enumerate([CourseDB.get_attendance(course_id, c_id, user['u_id']) for c_id in course['cw_numbers']]):
			sheet.write(i+3, j+1, attendance['value'])
		
	book.close()
	return file_name

def get_marks(course_id):
	file_name = 'ExcelData/Marks_{0}.xlsx'.format(course_id)
	book = xlsxwriter.Workbook(file_name)
	sheet = book.add_worksheet('Успеваемость')
	h_sheet = book.add_worksheet('Hidden')
	h_sheet.hide()

	course = CourseDB.get_course(course_id)

	centered_format = book.add_format(
									{
										'align': 'center',
                                   		'valign': 'vcenter',
                                   		'border': 1
                               		})
	users = list([UserDB.get_user(u_id) for u_id in course['participants']])
	sheet.merge_range(0, 0, 1, 0, 'Студенты', centered_format)
	sheet.set_column(0, 0, len(max([user['name'] for user in users])) + 4)

	task_count = len(course['task_numbers'])
	sheet.merge_range(0, 1, 1, 1+task_count, 'Задачи', centered_format)
	sheet.write(2, 1, 'Название')
	sheet.write(3, 1, 'Макс. балл')
	sheet.write(4, 1, 'Дедлайн')

	for i, task in enumerate([CourseDB.get_task(course_id, t_id) for t_id in course['task_numbers']]):
		sheet.write(2, 2+i, task['name'])
		h_sheet.write(2, 2+i, task['task_id'])
		sheet.write(3, 2+i, task['highest_mark'])
		sheet.write(4, 2+i, task['deadline'])

	for i, user in enumerate(users):
		sheet.write(i+5, 0, str(i+1) + '. ' + user['name'])
		h_sheet.write(i+5, 0, user['u_id'])

		for j, mark in enumerate([CourseDB.get_mark(course_id, t_id, user['u_id']) for t_id in course['task_numbers']]):
			sheet.write(i+5, j+2, mark['mark'])
		
	book.close()
	return file_name


def set_marks(course_id, excel_file):
	book = None
	try:
		book = openpyxl.load_workbook(excel_file)
	except:
		raise ExtensionException(excel_file)
	sheet = book.active

	p_file = get_marks(course_id)
	p_book = openpyxl.load_workbook(p_file)
	h_p_sheet = p_book['Hidden']
	p_sheet = p_book.active
	
	course = CourseDB.get_course(course_id)
	t_count = len(course['task_numbers'])
	u_count = len(course['participants'])

	for i in range(6, 6 + u_count):
		for j in range(3, 3 + t_count):
			int_cell = None
			try:
				int_cell = int(sheet.cell(i, j).value)
			except:
				p_book.close()
				book.close()
				raise CellValueException(i, j, '{0} must be an integer.'.format(sheet.cell(i, j).value))

			highest_mark = int(p_sheet.cell(4, j))
			if 0 <= int_cell <= highest_mark:
				u_id = h_p_sheet.cell(i, 1).value
				t_id = h_p_sheet.cell(3, j).value
				CourseDB.set_mark(course_id, t_id, u_id, 'value', str(int_cell))
			else:
				p_book.close()
				book.close()
				raise CellValueException(i, j, '{0} must be greater than 0 and less than {1}.'.format(int_cell, highest_mark))
	p_book.close()
	book.close()

def set_attendance(course_id, excel_file):
	book = None
	try:
		book = openpyxl.load_workbook(excel_file)
	except:
		raise ExtensionException(excel_file)
	sheet = book.active

	p_file = get_attendance(course_id)
	p_book = openpyxl.load_workbook(p_file)
	h_p_sheet = p_book['Hidden']
	p_sheet = p_book.active

	course = CourseDB.get_course(course_id)
	cw_count = len(course['cw_numbers'])
	u_count = len(course['participants'])

	for i in range(4, 4 + u_count):
		for j in range(2, 2 + cw_count):
			if str(sheet.cell(i, j).value) != '0' or str(sheet.cell(i, j).value) != '1':
				p_book.close()
				book.close()
				raise CellValueException(i, j, '{0} must be replaced with 0 or 1.'.format(sheet.cell(i, j).value))
			if str(sheet.cell(i, j).value) != str(p_sheet.cell(i, j).value):
				u_id = h_p_sheet.cell(i, 1).value
				cw_id = h_p_sheet.cell(3, j).value
				CourseDB.set_attendance(course_id, cw_id, u_id, 'value', str(sheet.cell(i, j).value))
	p_book.close()
	book.close()
